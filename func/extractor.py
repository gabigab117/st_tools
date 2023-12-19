import openpyxl

class Extractor:

    def __init__(self, file):
        self.data = self._open_doc(file)
        self.line = self.count_line()
        self.workbook = openpyxl.Workbook()

    def _open_doc(self, file):
        with open(file, "r") as f:
            data = f.read()
        return data
    
    def _clean_data(self):
        return self.data.replace("|", "").replace("-", "").replace("UVC", "")
    
    def _split_data(self):
        return self._clean_data().split("\n")
    
    def _groupe_line(self):
        return [data.split() for data in self._split_data()]
    
    def count_line(self):
        return self.data.count("OK")
    
    def get_ean_and_quantity(self):
        return [(el[0], el[1]) for el in self._groupe_line() if el]
    
    def save_in_file(self):
        with open("new.txt", "a") as file:
            for el in self.get_ean_and_quantity():
                file.write(f"{el[0]} {el[1]}\n")
    
    def _init_workbook(self):
        sheet = self.workbook.active
        cell = sheet.cell(row=1, column=1)
        cell_2 = sheet.cell(row=1, column=2)
        cell_3 = sheet.cell(row=1, column=3)
        cell.value = "EAN"
        cell_2.value = "Quantité commandée"
        cell_3.value = f"{self.line} lignes"

    
    def save_in_workbook(self):
        self._init_workbook()
        sheet = self.workbook.active

        for i, el in enumerate(self.get_ean_and_quantity(), 2):
            cell = sheet.cell(row=i, column=1)
            cell_2 = sheet.cell(row=i, column=2)
            cell.value = int(el[0])
            cell_2.value = int(el[1])
        
        self.workbook.save("commande.xlsx")


class TextAreaExtractor(Extractor):
    def __init__(self, file):
        self.data = file
        self.line = self.count_line()
        self.workbook = openpyxl.Workbook()

    

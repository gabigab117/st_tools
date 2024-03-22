import openpyxl
from difflib import get_close_matches
from uuid import uuid4
from datetime import datetime
import pathlib

class Matchinator:
    def __init__(self, rejects, orderable, user: str):
        self.rejects = self.rejects_sheet(rejects)
        self.orderable = self.orderable_sheet(orderable)
        self.rejects_dict = self.create_rejects_dict()
        self.orderable_dict = self.create_orderable_dict()
        self.match_dict = self.create_match_dict()
        self.date = datetime.now()
        self.user = user
    
    def rejects_sheet(self, rejects):
        """Load and return the rejects sheet from the rejects workbook."""
        return openpyxl.load_workbook(rejects).worksheets[0]
    
    def orderable_sheet(self, orderable):
        """Load and return the orderable sheet from the orderable workbook."""
        return openpyxl.load_workbook(orderable).worksheets[0]
    
    def create_rejects_dict(self):
        """Create a dictionary from rejects sheet mapping product codes to descriptions."""
        rejects = {}
        for cel_d, cell_f in zip(self.rejects["D"], self.rejects["F"]):
            rejects[cel_d.value] = cell_f.value
    
        return rejects
    
    def create_orderable_dict(self):
        """Create a dictionary from orderable sheet mapping product codes to descriptions."""
        cadencier = {}
        for cel_l, cell_b in zip(self.orderable["L"], self.orderable["B"]):
            cadencier[cel_l.value] = cell_b.value
        return cadencier
    
    def create_match_dict(self):
        """Create a dictionary mapping rejected products to matching orderable products."""
        reach = {}
        for v in self.rejects_dict.values():
            matches = get_close_matches(v, self.orderable_dict.values())
            reach[v] = {
                match_item: next(key for key, value in self.orderable_dict.items() if value == match_item)
                for match_item in matches
            }
        return reach
    
    def create_match_workbook(self):
        """Create and save a workbook containing matched rejected and orderable products."""
        new_wb = openpyxl.Workbook()
        active_sheet = new_wb.active
        row_index = 1
        for reject_product, dict_interieur in self.match_dict.items():
            active_sheet.cell(row=row_index, column=1, value=reject_product)
            row_index += 1
    
            for (product, code) in dict_interieur.items():
                active_sheet.cell(row=row_index, column=2, value=code)
                active_sheet.cell(row=row_index, column=3, value=product)
                row_index += 1
        
        folder = pathlib.Path("match")
        folder.mkdir(exist_ok=True)
        random_str = uuid4()
        file_name = f"match{random_str}_{self.date.day}_{self.date.month}_{self.date.year}_{self.user}.xlsx"
        file_path = str(folder/ file_name)

        new_wb.save(file_path)
        
        return file_path

